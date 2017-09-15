#!/usr/bin/env python2.7
from __future__ import division
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
import os
from datetime import datetime 
import pyperclip
import requests
import api_google
from mortgage import Mortgage

from config import CATEGORIES, FILE_NAME, SHEET_NAME, SHEET_MLS_INDEX_NUM, MORTGAGE_INTEREST_RATE, MORTGAGE_LOAN_YEARS, ENABLE_CLOUD_SAVING, MORTGAGE_DOWNPAYMENT_PERCENTAGE, IS_CLOUD_SOURCE_OF_TRUE

def strip_number(num):
    return round(float(num.replace('$', '').replace(',','')), 2)

def update_data_with_extra_columns(row, currRow): 
    taxesPerMonth = strip_number(row['taxes'])
    listPrice = strip_number(row['listPr'])
    mortgageAmount = listPrice * (1 - (MORTGAGE_DOWNPAYMENT_PERCENTAGE/100))
    downpayment = listPrice * MORTGAGE_DOWNPAYMENT_PERCENTAGE / 100
    mortgageMonthly = Mortgage(float(MORTGAGE_INTEREST_RATE) / 100, MORTGAGE_LOAN_YEARS * 12, mortgageAmount)

    row.update({'#': currRow})
    row.update({'dateAdded': datetime.now().strftime('%Y-%m-%d')})
    row.update({'taxesPerMonth': round(taxesPerMonth/12,2)})
    row.update({'estMonthlyMortgage': mortgageMonthly.monthly_payment()})
    row.update({'downPayment': round(downpayment, 2)})
    row.update({'mortgageAmt': round(mortgageAmount, 2)})
    return row

def remove_duplicates(rowsToSave):
    seen = set()
    new_l = []
    for d in rowsToSave:
        t = tuple(d.items())
        if t not in seen:
            seen.add(t)
            new_l.append(d) 
    return new_l

def create_headers_for_sheet(sheet):
    for index, elem in enumerate(CATEGORIES):
        sheet.cell(row=1,column=index+1).value=elem
    return sheet

def get_excel_file():
    if os.path.isfile(FILE_NAME):
        print 'retrieving existing file'
        wb = load_workbook(FILE_NAME)
        return wb
    else:
        print 'creating new file'
        wb = Workbook()
        wb.save(FILE_NAME)
        return wb


def get_excel_sheet(wb):
    sheets = wb.sheetnames
    if SHEET_NAME not in sheets:
        print SHEET_NAME + ' sheet not found. Creating new sheet...'
        sheet = wb.create_sheet(SHEET_NAME, 0)
        create_headers_for_sheet(sheet)
        wb.save(FILE_NAME)
    return wb.get_sheet_by_name(SHEET_NAME)

def write_data_to_excel(rowsToSave, sheet): 
    rowToStart = sheet.max_row + 1
    # TODO: Check if exists
    print 'Saving data...'
    listingsAdded = 0
    currRow = rowToStart
    for row in rowsToSave:
        # Check if MLS number already exists
        listingAlreadyExists = False
        for currRowCheck in range(2, rowToStart):
            if sheet.cell(row=currRowCheck, column=SHEET_MLS_INDEX_NUM).value == row['mlsNum']:
                print 'Listing ' + str(row['mlsNum']) + ' already exists. Skipping...'
                listingAlreadyExists = True
                break
        
        if listingAlreadyExists:
            continue
        row = update_data_with_extra_columns(row, currRow)

        i = 1
        for val in CATEGORIES:
            sheet.cell(row=currRow,column=i, value=row[val])
            i+=1
        currRow+=1
        listingsAdded+=1
    print str(listingsAdded) + ' new listings added...'
    return listingsAdded
    

def do_excel_stuff(data):
    wb = get_excel_file()
    sheet = get_excel_sheet(wb)
    listingAdded = write_data_to_excel(data, sheet)
    wb.save(FILE_NAME)
    return listingAdded

def extractTextData(textData):
    av = textData.select('span.value')
    textDescriptions = ''
    for v in av:
        textDescriptions+=v.get_text()
    return {
        'textDescriptions' : textDescriptions
    }


def extractRoomSizeData(roomSizeData):
    return
    # av = roomSizeData.select('span.value')

def extractSpecificData(specificData):
    av = specificData.select('span.value')
    return {
        'numKitchens': av[0].get_text(),
        'hasBasement': av[2].get_text(),
        'hasFireplace': av[3].get_text(),
        'heatType': av[4].get_text(),
        'approxAge': av[5].get_text(),
        'approxSqFeet': av[6].get_text(),
        'sqftSource': av[7].get_text(),
        'unitDirection': av[8].get_text(),
        'pets':av[12].get_text(),
        'locker':av[13].get_text(),
        'maintenance':av[14].get_text(),
        'airCon':av[15].get_text(),
        'taxesIncl':av[20].get_text(),
        'waterIncl':av[21].get_text(),
        'heatIncl':av[22].get_text(),
        'hydroIncl':av[23].get_text(),
        'cablTvIncl':av[24].get_text(),
        'centralAirCon':av[25].get_text(),
        'buildingInsurIncl':av[26].get_text(),
        'parkingIncl':av[27].get_text(),
        'balconyType':av[31].get_text(),
        'ensuiteLaundry':av[32].get_text(),
        'exterior':av[34].get_text(),
        'garage':av[35].get_text(),
        'parking':av[36].get_text(),
        'parkType':av[37].get_text(),
        'parkSpaces':av[38].get_text(),
        'totalParkSpots':av[39].get_text(),
        'parkSpotNum':av[40].get_text(),
        'parkPerMonth':av[41].get_text(),
        'parkLevel':av[42].get_text(),
        'commonElemIncl':av[43].get_text()
    }

def extractMiscData(miscData): 
    av = miscData.select('span.value')
    mlsNum = av[0].get_text()
    posessionDate = av[1].get_text()
    return {
        'mlsNum': mlsNum,
        'possessionDate': posessionDate
    }

def extractTopLevelData(topLevelData):
    av = topLevelData.select('span.value')
    return  {
        'address': av[0].get_text(),
        'unit': av[1].get_text(),
        'city': av[2].get_text(),
        'prov': av[3].get_text(),
        'postal': av[4].get_text(),
        'listPr': av[5].get_text(),
        'salesStatus': av[6].get_text(),
        'taxes': av[12].get_text(),
        'taxYear': av[13].get_text(),
        'lastStatus': av[14].get_text(),
        'unitType': av[15].get_text(),
        'unitType2': av[16].get_text(),
        'lockerNum': av[19].get_text(),
        'lockerLevel': av[20].get_text(),
        'lockerUnit': av[21].get_text(),
        'floorNum': av[22].get_text(),
        'unitNum': av[23].get_text(),
        'roomCount': av[24].get_text(),
        'bedroomCount': av[25].get_text(),
        'washroomCount': av[26].get_text(),
        'washroomTypes': av[27].get_text(),
        'crossSt': av[28].get_text(),
    }


def print_app_mode():
    print '-------SETTINGS------'
    if ENABLE_CLOUD_SAVING:
        print '[Settings] Google Drive Saving Enabled.'
        if IS_CLOUD_SOURCE_OF_TRUE:
            print '[Settings] Cloud files are source of truth'
        else:
            print '[Settings] Local File is source of truth.'
    else:
        print '[Settings] Local Saving Enabled.'
    print '------END OF SETTINGS------'

def main(): 
    print_app_mode()
    if ENABLE_CLOUD_SAVING:
        file_resource = api_google.get_latest_file_from_drive()
    url = raw_input('enter a site: \n')
    data = requests.get(url)
    html_content = data.text
    soup = BeautifulSoup(html_content, 'html.parser')

    pyperclip.copy(soup.prettify())

    high_level_listings = soup.html.find_all('div', attrs={'class': 'report-container'}) or []

    a_listings = soup.html.find_next_siblings('div', attrs={"class": "report-container"}) or []
    b_listings = soup.html.find_next_siblings('div', attrs={"class": "link-item"}) or []

    for l in a_listings:
        b_listings.append(l)

    for x in high_level_listings:
        b_listings.append(x)

    rowsToSave = []
    for data in b_listings: 
        data = data.find('div', attrs={"class": "legacyBorder"})

        if data is None: continue

        # top level info
        datasets = data.select('> .formgroup.formitem')

        topLevelData = datasets[0]
        miscData = datasets[1]
        specificData = datasets[2]
        roomSizeData = datasets[5]
        textData = datasets[7]


        dataToSave = {}

        topDataDict = extractTopLevelData(topLevelData)
        miscDataDict = extractMiscData(miscData)
        specificDataDict = extractSpecificData(specificData)
        # roomSizeDict = extractRoomSizeData(roomSizeData)
        textDataDict = extractTextData(textData)

        dataToSave.update(topDataDict)
        dataToSave.update(miscDataDict)
        dataToSave.update(specificDataDict)
        dataToSave.update(textDataDict)

        rowsToSave.append(dataToSave)    

    rowsToSave = remove_duplicates(rowsToSave)
    listingAdded = do_excel_stuff(rowsToSave)
    if listingAdded and ENABLE_CLOUD_SAVING:
        api_google.saveIntoGoogleDrive(file_resource)
        

if __name__ == '__main__':
    main()